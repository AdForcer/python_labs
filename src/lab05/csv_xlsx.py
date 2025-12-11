# src/lab05/csv_xlsx.py
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:

    csv_file = Path(csv_path)
    xlsx_file = Path(xlsx_path)

    # Проверка существования файла
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    # Чтение CSV данных
    rows = []
    try:
        with csv_file.open("r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)
    except csv.Error as e:
        raise ValueError(f"Ошибка чтения CSV: {e}")

    # Валидация данных
    if not rows:
        raise ValueError("CSV файл пустой")

    # Создание Excel файла
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Запись данных
    for row in rows:
        ws.append(row)

    # Настройка авто-ширины колонок
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Минимальная ширина - 8 символов
        adjusted_width = max(max_length + 2, 8)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохранение файла
    xlsx_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_file)
